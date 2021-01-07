import pandas
from openpyxl import load_workbook
from pandas import DataFrame

wb = load_workbook(r'D:\DESARROLLO\busca_dni\_mensajes_2021-01-07113221.xlsx')
ws = wb.active

wb_b = load_workbook(r'D:\DESARROLLO\busca_dni\LISTADO SEGUNDAS CUOTAS CAPITAL.xlsx')
ws_b = wb_b.active

beneficiarios = []
consultas = []
for i, row in enumerate (ws_b.values, start=1):
	beneficiarios.append([str(row[2]), row[16]])
pass
for i, row in enumerate (ws.values, start=1):
	consultas.append([str(row[3]), str(i)])
pass
dni_consultas = []
for consulta in consultas:
    dni=[str(Numero) for Numero in consulta[0].split() if Numero.isdigit() and len(Numero)==8 or Numero.isdigit() and len(Numero)==7]
    if bool(dni) == True:
        dni_consultas.append([dni[0],consulta[1]])

pass
df_benef = DataFrame (beneficiarios,columns=['DNI', 'estado'])
df_dnicons = DataFrame (dni_consultas,columns=['DNI', 'row'])
df = df_dnicons.merge(df_benef, how="inner")
print(df)

for i in df.index: 
    ws['E'+str(df.at[i, 'row'])] = str(df.at[i, 'estado'])
pass
wb.save(r'D:\DESARROLLO\busca_dni\parseo_cruzado.xlsx')