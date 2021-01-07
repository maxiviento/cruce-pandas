import pandas
from openpyxl import load_workbook
from pandas import DataFrame

#CARGA BENEF DE GOOGLE DRIVE
key = '1HODwblDChV6NNKvrAp2j7Bj4u37id8p7l_a9IB9GavQ'
hoja = 'Sheet'
url = 'https://docs.google.com/spreadsheets/d/{}/gviz/tq?tqx=out:csv&sheet={}'.format(key,hoja)
df_benef = pandas.read_csv(url, usecols=['NRO_DOCUMENTO','REPORTE DE LLAMADO'])

df_benef = df_benef.rename(columns={'NRO_DOCUMENTO':'DNI'})
df_benef[['DNI']] = df_benef[['DNI']].astype('object', copy=False)


#CARGA PARSEO DE GOOGLE DRIVE
wb = load_workbook(r'D:\DESARROLLO\busca_dni\_mensajes_2021-01-06153610.xlsx')
ws = wb.active

consultas = []

for i, row in enumerate (ws.values, start=1):
	consultas.append([str(row[3]), str(i)])
pass
dni_consultas = []
for consulta in consultas:
    dni=[str(Numero) for Numero in consulta[0].split() if Numero.isdigit() and len(Numero)==8 or Numero.isdigit() and len(Numero)==7]
    if bool(dni) == True:
        dni_consultas.append([dni[0],consulta[1]])

pass
df_dnicons = DataFrame (dni_consultas,columns=['DNI', 'row'])
df = df_dnicons.merge(df_benef, how="inner")
print(df_benef)
print(df_dnicons)
#print(df)

for i in df.index: 
    print(str(df.at[i, 'row']) + ", " + str(df.at[i, 'estado']))
    ws['E'+str(df.at[i, 'row'])] = str(df.at[i, 'estado'])
pass
wb.save(r'D:\DESARROLLO\busca_dni\parseo_cruzado.xlsx')





